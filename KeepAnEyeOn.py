import DataFactory as dfy
import WebCrawler as wc
import openpyxl as xl
import datetime as dt
from openpyxl.styles import colors, Font
import os
import time

class KeepAnEyeOn:

    def __init__(self):
        self.priordata = {}
        self.prior_columns = 0
        self.today_exl = {}
        self.is_exception = 0

    def kaeoWorker(self):
        root_path = 'C:\\MyProjects\\output\\'
        excel_path = root_path + dt.date.today().isoformat() + '.xlsx'


        for prior_day in range(5):
            prior_excel_path = root_path + (dt.date.today() - dt.timedelta(days=prior_day+1)).isoformat() +'.xlsx'
            if os.path.isfile(prior_excel_path):
                self.readexcel(prior_excel_path)
                break

        if self.is_exception == 0:
            self.todaypricefetcher()
            self.writeexcel(excel_path)


    # read data from existing sheet, and write them into self.priordata
    def readexcel (self, path):

        workbook = xl.load_workbook(path)

        try:
            sheet = workbook['KAEO']
            self.prior_columns = sheet.max_column

            if self.prior_columns == 0:
                return

            # exclude first row (titles)
            if sheet._get_cell(row=1, column=1).value == 'symbol':
                sheet.delete_rows(idx=1, amount=1)

            # has 10 days of data, need to get rid of the first day
            if self.prior_columns == 12:
                sheet.delete_columns(idx=3, amount=1)

            # put all data into list, convert None to ''
            g = lambda x: x if x is not None else ''
            for row in sheet.rows:
                list_temp = []
                for cell in row:
                    list_temp.append(g(cell.value))
                self.today_exl[list_temp[0].upper()] = list_temp[1:]
        except Exception as e:
            print('readexcel: exception when reading the excel')
            print(e)
            self.is_exception =1



    # get a list of symbols, and call webcrawler.py, put them in self.todaypricelist
    def todaypricefetcher (self):

        webc = wc.WebCrawler()

        # go through each symbols in the kaeo
        for eachSymbol in self.today_exl.keys():

            print('KAEO: processing ' + eachSymbol)
            symboldata_url = 'https://query1.finance.yahoo.com/v8/finance/chart/' + eachSymbol + '?&region=US&lang=en-US&includePrePost=false&interval=1d&range=2d&corsDomain=finance.yahoo.com'
            df = dfy.DataFactory()
            df.json_digest(webc.url_open(symboldata_url))
            if df.is_excepttion == 0:
                if df.close_price != 0:

                    # compute the change %:
                    change = ((df.close_price[-1]-df.close_price[0])/df.close_price[0])*100
                    todaycell = '[' + '%.2f' % change + '%_' + '%.2f' % df.close_price[-1] +'<' + \
                                str('%.1f' % (df.volume[-1] / 1000000)) + 'M>'

                    self.today_exl[eachSymbol].append(todaycell)


            time.sleep(2)

    # write them in today excel
    def writeexcel (self, excel_path):
        # if there is excel, add a new sheet to it, if not, create a new excel:
        if os.path.isfile(excel_path):
            workbook = xl.load_workbook(excel_path)
        else:
            workbook = xl.Workbook()

        sheet = workbook.create_sheet('KAEO')

        titles = ['symbol', 'note']
        iter_number = 0

        # write the title row:
        if self.prior_columns == 12:
            iter_number = self.prior_columns - 2
        elif self.prior_columns > 0 and self.prior_columns < 12:
            iter_number = self.prior_columns - 1
        elif self.prior_columns == 0:
            iter_number = 1

        for i in range(iter_number):
            titles.append(i + 1)

        for each in range(len(titles)):
            sheet.cell(row=1, column=each + 1, value=str(titles[each]))

        # load symbols in list, and sort them --- should I maintain the sorting myself???
        symbol_list = []
        for each in self.today_exl.keys():
            symbol_list.append(each)
        symbol_list.sort()

        rf = Font(color=colors.RED)

        # color link: http://www.114la.com/other/rgb.htm
        # 00218868 = #218868
        gf = Font(color="00218868")

        for each in range(len(symbol_list)):
            sheet.cell(row=each + 2, column=1, value=symbol_list[each])

            for lst in range(len(self.today_exl[symbol_list[each]])):
                sheet.cell(row=each + 2, column=lst + 2, value=str(self.today_exl[symbol_list[each]][lst]))

                # skip the note colume, set color red/green for the font in cells based on the down/up day
                if lst != 0:
                    if self.today_exl[symbol_list[each]][lst].find('-') == -1:
                        sheet.cell(row=each + 2, column=lst + 2).font = gf
                    else:
                        sheet.cell(row=each + 2, column=lst + 2).font = rf

        workbook.save(excel_path)



if __name__ == '__main__':
    keepeye = KeepAnEyeOn()
    keepeye.kaeoWorker()