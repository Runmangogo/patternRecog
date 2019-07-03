import openpyxl as xl
import datetime as dt
from openpyxl.styles import PatternFill
import os

class merge2Excel:

    def __init__(self):
        self.prior_columns = 0
        self.prior_exl_data = {}
        self.today_data = {}
        self.today_exl = {}


    def merge2ExcelWorker (self, index, list_type):
        # list_type = '_HighAlert'
        root_path = 'C:\\MyProjects\\output\\'
        list_path = root_path + dt.date.today().isoformat() + '_' + index + list_type + '.txt'
        #excel_path = root_path + dt.date.today().isoformat() + '_' + index + list_type +  '.xlsx'
        #prior_excel_path = root_path + (dt.date.today()-dt.timedelta(days=1)).isoformat() + '_' + index + list_type +  '.xlsx'
        excel_path = root_path + dt.date.today().isoformat() + '.xlsx'

        sheet_name = index + list_type


        # check if the file exists

        # go back 5 days to make sure it can get the prior xlsx (when in between there is weekend, long weekend, holiday, etc,)
        for prior_day in range(5):
            prior_excel_path = root_path + (dt.date.today() - dt.timedelta(days=prior_day+1)).isoformat() +'.xlsx'
            if os.path.isfile(prior_excel_path):
                self.read_excel_xlsx(prior_excel_path, sheet_name)
                break



        if os.path.isfile(list_path) == False :
            print('file not found: ' + list_path)
            return

        self.read_txt(list_path)

        # need to do compare two sets of data: prior data in excel vs today's data, then merge then sort
        self.dataMassage()

        # write the data into new excel file
        self.write_excel_xlsx(sheet_name, excel_path)

        print('saved data in new xlsx:' + index + list_type)


    # read prior excel into cache:
    def read_excel_xlsx(self, path, sheet_name):
        workbook = xl.load_workbook(path)
        sheet = workbook[sheet_name]
        self.prior_columns = sheet.max_column

        if self.prior_columns == 0:
            return

        # exclude first row (titles)
        if sheet._get_cell(row=1, column=1).value == 'symbol':
            sheet.delete_rows(idx=1, amount=1)

        # has 10 days of data, need to get rid of the first day
        if self.prior_columns == 12:
            sheet.delete_columns(idx=3, amount=1)

        g = lambda x: x if x is not None else ''
        for row in sheet.rows:
            list_temp = []
            for cell in row:
                list_temp.append(g(cell.value))

            # old codes for : has 10 days of data, need to get rid of the first day
            # if self.prior_columns == 12:
            #    list_temp.remove(list_temp[2])

            # for each row data in the list, convert them to dict
            # bug fix: only add the symbol with value in any of days -- do not add symbol with None value in all days
            hasvalue = 1
            # I dont like codes below, improve it in the future.
            if self.prior_columns == 12:
                hasvalue = 0
                for element in list_temp[2:]:
                    if element != '':
                        hasvalue = 1
                        break
            if hasvalue == 1:
                self.today_exl[list_temp[0].upper()] = list_temp[1:]

        # debug
        # print('today_exl(before merge):' + str(self.today_exl))

    # read txt file into cache:
    def read_txt(self, list_path):
        f = open(list_path, 'r')
        read_txt = f.readlines()

        if read_txt == None:
            return

        for each in read_txt:
            self.today_data[each[:each.find('|')]] = each[each.find('|') + 1:].replace('\n', '')

        # debug:
        # print('today_data(from txt):' + str(self.today_data))

    # compare two sets of data, and merge and re_sort them
    def dataMassage(self):

        for each in self.today_data.keys():

        # debug
        # print(self.prior_columns)

            # if it's a new symbol
            if self.today_exl.get(each) == None:

                fill_None = 0
                # add None columns to today_data[each], to fit the number of columns
                if self.prior_columns == 12:
                    fill_None = self.prior_columns - 2
                elif self.prior_columns > 0 and self.prior_columns < 12:
                    fill_None = self.prior_columns - 1
                elif  self.prior_columns == 0:
                    fill_None = 1


                self.today_exl[each] = []
                for each_None in range(fill_None):
                    self.today_exl[each].append('')

            # add today's data to the end of the list
            self.today_exl[each].append(self.today_data[each])

        # debug
        # print('today_exl(after merge):' + str(self.today_exl))


    def write_excel_xlsx(self, sheet_name, excel_path):
        #index = len(value)

        # if there is excel, add a new sheet to it, if not, create a new excel:
        if os.path.isfile(excel_path):
            workbook = xl.load_workbook(excel_path)
        else:
            workbook = xl.Workbook()

        sheet = workbook.create_sheet(sheet_name)

        titles = ['symbol', 'note']
        iter_number = 0

        # write the title row:
        if self.prior_columns == 12:
            iter_number = self.prior_columns - 2
        elif self.prior_columns > 0 and self.prior_columns < 12:
            iter_number = self.prior_columns - 1
        elif self.prior_columns == 0:
            iter_number = 1

        for i in range(iter_number):
            titles.append(i+1)

        for each in range(len(titles)):
            sheet.cell(row=1, column=each+1, value=str(titles[each]))


        # load symbols in list, and sort them
        symbol_list = []
        for each in self.today_exl.keys():
            symbol_list.append(each)
        symbol_list.sort()

        # initiate orange colors #FFC125 for cell
        fill = PatternFill("solid", fgColor="00FFC125")

        # write data to excel
        for each in range(len(symbol_list)):
            sheet.cell(row=each + 2, column=1, value= symbol_list[each])
            #self.today_exl[symbol_list[each]].replace(None, '')
            for lst in range(len(self.today_exl[symbol_list[each]])):
                sheet.cell(row=each + 2, column=lst + 2, value=str(self.today_exl[symbol_list[each]][lst]))

            # fill the row of cells if the symbol match the pattern I want
            if self.today_exl[symbol_list[each]][0].casefold().find('>match<') != -1:
                for lst in range(len(self.today_exl[symbol_list[each]])):
                    sheet.cell(row=each + 2, column=lst + 2).fill = fill

        workbook.save(excel_path)





#write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
#read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)


if __name__ == '__main__':
    mer = merge2Excel()
    mer.merge2ExcelWorker('RussellMidCap', '_HighAlert')
    mer2 = merge2Excel()
    mer2.merge2ExcelWorker('Russell2000', '_HighAlert')
